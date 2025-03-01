import openpyxl as xl
from openpyxl.chart import Reference, BarChart

def process_workbook(filename):
    wb = xl.load_workbook('../transactions.xlsx')
    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell = sheet.cell(1,1)

    for row in range(2, sheet.max_row + 1):
        # print(row)
        cell = sheet.cell(row, 3)
        # print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save('transactions.xlsx')